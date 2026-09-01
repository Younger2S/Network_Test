# -*- coding: utf-8 -*-
"""ES 数据库弱口令检测工具 - Flask 后端

功能说明：
1. 文件上传：支持 .xlsx（指定列含 ES URL/IP）与 .txt（默认取第 1 列），含格式校验；
2. 参数选择：URL 所在列数仅允许 1-999 的数字，指定列数超过表格实际列数时给出提示；
3. 目标解析：支持 http(s)://host:port、host:port、纯 IPv4、域名，自动识别端口；
4. 结果输出：列表展示检测结果，支持单条下载、全部下载、删除、清空。
"""
import io
import os
import re
import json
import uuid
import logging
import ipaddress
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlsplit

from flask import Flask, jsonify, render_template, request, send_file

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    requests = None
    REQUESTS_OK = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    Workbook = load_workbook = Font = get_column_letter = None
    OPENPYXL_OK = False

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
BATCH_DIR = os.path.join(BASE_DIR, "batches")   # 上传批次上下文（用于还原输入表格）
RESULTS_FILE = os.path.join(BASE_DIR, "results.json")
MAX_FILE_SIZE = 10 * 1024 * 1024          # 上传文件大小上限：10MB
ALLOWED_EXTENSIONS = {".xlsx", ".txt"}    # 扩展名白名单
DEFAULT_PORT = 9200                        # ES 默认端口
MAX_THREADS = 10                           # 并发检测线程数

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(BATCH_DIR, exist_ok=True)

# ==================== 日志 ====================
LOG_DIR = os.path.join(BASE_DIR, "log")


class DateFileHandler(logging.Handler):
    """按日期滚动写日志：log/yyyyMMdd.log，跨天自动切换文件。"""

    def __init__(self, log_dir):
        super().__init__()
        self.log_dir = log_dir
        self._date = None
        self._stream = None

    def _stream_for(self):
        today = datetime.now().strftime("%Y%m%d")
        if self._date != today:
            if self._stream:
                self._stream.close()
            os.makedirs(self.log_dir, exist_ok=True)
            self._stream = open(os.path.join(self.log_dir, f"{today}.log"), "a", encoding="utf-8")
            self._date = today
        return self._stream

    def emit(self, record):
        try:
            stream = self._stream_for()
            stream.write(self.format(record) + "\n")
            stream.flush()
        except Exception:
            pass


os.makedirs(LOG_DIR, exist_ok=True)
logger = logging.getLogger("es_checker")
logger.setLevel(logging.INFO)
if not logger.handlers:
    _fh = DateFileHandler(LOG_DIR)
    _fh.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    logger.addHandler(_fh)
    _ch = logging.StreamHandler()
    _ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(_ch)

# ==================== 弱口令检测规则 ====================
# 常规模式：精选常见默认口令与典型弱口令组合
WEAK_PASSWORDS = [
    ("elastic", "changeme"),
    ("elastic", "elastic"),
    ("elastic", "123456"),
    ("elastic", "password"),
    ("elastic", "es123456"),
    ("elastic", "1qaz2wsx"),
    ("admin", "admin"),
    ("admin", "123456"),
    ("admin", "changeme"),
    ("admin", "password"),
    ("admin", "admin123"),
    ("root", "root"),
    ("root", "123456"),
    ("root", "password"),
    ("es_user", "12345678"),
    ("es_user", "es_user"),
    ("kibana", "kibana"),
    ("kibana", "kibana123"),
    ("logstash", "logstash"),
    ("system", "manager"),
    ("guest", "guest"),
    ("apm_system", "apm_system"),
]

# 深度检测（爆破式）用户名表：ES 官方内置用户 + 常见运维/默认账号
DEEP_USERNAMES = [
    "elastic", "admin", "root", "es_user", "kibana", "logstash",
    "beats", "apm_system", "remote_monitoring_user", "guest",
    "system", "read_user",
    "kibana_system", "logstash_system", "elastic_system", "beats_system",
    "monitoring_user", "superuser", "viewer", "editor",
    "kibana_admin", "watcher_admin", "ml_admin", "ingest_admin",
    "reporting_user", "transport_client", "test", "user",
    "es_admin", "elk", "elasticsearch",
]

# 深度检测密码规则表：纯数字 / 全小写单词 / 键盘序列 / 单词+数字 / 服务名衍生
DEEP_PASSWORD_WORDS = [
    # 纯数字
    "123456", "12345678", "888888", "666666", "111111", "000000",
    "112233", "123123", "987654", "123321", "654321", "520520",
    "2024", "2023", "2022", "2020", "1990",
    # 全小写常用弱口令
    "password", "changeme", "secret", "letmein", "welcome",
    "admin", "root", "test", "guest", "user", "default", "public",
    "manager", "oracle", "mysql", "postgres", "elastic",
    "kibana", "logstash", "beats", "monitor", "espass",
    # 键盘相邻序列
    "qwerty", "qwe123", "1qaz2wsx", "zxcvbnm", "asdfgh",
    "zaq1zaq1", "qazwsx", "1q2w3e4r", "asdf1234",
    # 单词 + 数字 / 特殊字符变体
    "password1", "password123", "admin888", "root123", "test123",
    "abc123", "admin@123", "elastic123", "es@123", "es123456",
]


def _deep_credentials() -> list:
    """规则化生成深度检测的全部 (用户名, 密码) 组合（去重，保持顺序）。

    组合方式：常规模式组合 + 每个用户名派生"用户名本身 / 用户名+数字 / 用户名@123"，
    再与深度密码规则表交叉，覆盖全小写、纯数字、短口令、键盘序列等弱口令规则。
    """
    creds = list(WEAK_PASSWORDS)
    for u in DEEP_USERNAMES:
        creds.append((u, u))
        creds.append((u, u + "123"))
        creds.append((u, u + "123456"))
        creds.append((u, u + "@123"))
        creds.append((u, u + "888"))
        for p in DEEP_PASSWORD_WORDS:
            creds.append((u, p))
    seen, out = set(), []
    for c in creds:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out


DEEP_CREDENTIALS = _deep_credentials()   # 模块加载时生成一次，供深度检测复用

IP_RE = re.compile(r"^\s*(\d{1,3}(?:\.\d{1,3}){3})\s*$")
HOST_RE = re.compile(
    r"^[A-Za-z0-9](?:[A-Za-z0-9\-]{0,61}[A-Za-z0-9])?"
    r"(?:\.[A-Za-z0-9](?:[A-Za-z0-9\-]{0,61}[A-Za-z0-9])?)*$"
)


def is_valid_ip(ip: str) -> bool:
    """校验是否为合法的 IPv4 地址。"""
    try:
        return ipaddress.ip_address(ip).version == 4
    except ValueError:
        return False


def is_valid_host(host: str) -> bool:
    """校验是否为合法的域名/主机名。"""
    return bool(HOST_RE.match(host))


def validate_column(value) -> int:
    """校验 URL 所在列数：仅允许 1-999 之间的数字，否则返回 None。"""
    if value is None:
        return None
    s = str(value).strip()
    if not s or not s.isdigit():
        return None
    col = int(s)
    return col if 1 <= col <= 999 else None


def parse_target(raw) -> dict:
    """解析 ES URL/IP 字符串，返回 {target, host, port, scheme}；无法解析返回 None。

    支持格式：
        http://127.0.0.1:9201 / https://es.example.com:9200
        127.0.0.1:9200 / es.example.com:9200
        127.0.0.1 / es.example.com（端口/协议为空，检测时自动尝试 http/https）
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None

    # 带协议前缀的 URL
    if "://" in s:
        try:
            parts = urlsplit(s)
            host = parts.hostname
            if not host or parts.scheme not in ("http", "https"):
                return None
            try:
                port = parts.port
            except ValueError:
                return None
            return {"target": s, "host": host, "port": port, "scheme": parts.scheme}
        except ValueError:
            return None

    # 形如 host:port 或 ip:port
    m = re.match(r"^(.+?):(\d{1,5})$", s)
    if m:
        host, port_s = m.group(1), m.group(2)
        port = int(port_s)
        if 1 <= port <= 65535 and (is_valid_ip(host) or is_valid_host(host)):
            return {"target": s, "host": host, "port": port, "scheme": None}
        return None

    # 纯 IPv4
    if is_valid_ip(s):
        return {"target": s, "host": s, "port": None, "scheme": None}

    # 纯域名
    if is_valid_host(s):
        return {"target": s, "host": s, "port": None, "scheme": None}

    return None


def safe_remove(path: str) -> None:
    """安全删除文件，忽略删除失败。"""
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


def load_results() -> list:
    """读取本地持久化的检测记录。"""
    if not os.path.exists(RESULTS_FILE):
        return []
    try:
        with open(RESULTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except (json.JSONDecodeError, OSError):
        return []


def save_results(results: list) -> None:
    """持久化检测记录。"""
    with open(RESULTS_FILE, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


def decode_bytes(raw: bytes) -> str:
    """按优先级尝试多种编码解码字节流。"""
    for enc in ("utf-8", "gbk", "gb18030", "latin-1"):
        try:
            return raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("utf-8", errors="ignore")


def read_xlsx_rows(path: str) -> tuple:
    """读取 xlsx 全部单元格，返回 (rows, max_cols)。

    仅读取表格结构，不在此处校验 URL 列（URL 校验推迟到点击"开始查询"时）。
    """
    if not OPENPYXL_OK:
        raise RuntimeError("缺少 openpyxl 依赖，请先执行 pip install -r requirements.txt")
    wb = load_workbook(path, read_only=True, data_only=True)
    rows, max_cols = [], 0
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals = ["" if v is None else v for v in row]
                rows.append(vals)
                max_cols = max(max_cols, len(vals))
    finally:
        wb.close()
    return rows, max_cols


def extract_targets_from_rows(rows: list, column: int) -> tuple:
    """从已读取的表格行按指定列提取 ES URL/IP。

    返回 (targets, target_rows)：target_rows 为目标字符串 -> 1 起始行号。
    """
    targets, seen, target_rows = [], set(), {}
    for row_no, vals in enumerate(rows, start=1):
        if column <= len(vals):
            t = parse_target(vals[column - 1])
            if t and t["target"] not in seen:
                seen.add(t["target"])
                targets.append(t)
                target_rows[t["target"]] = row_no
    return targets, target_rows


def read_txt_lines(path: str) -> list:
    """读取 txt 文本的非空有效行（跳过以 # 或 // 开头的注释行）。"""
    with open(path, "rb") as f:
        raw = f.read()
    text = decode_bytes(raw)
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if line and not line.startswith(("#", "//")):
            lines.append(line)
    return lines


def extract_targets_from_lines(lines: list, column: int) -> tuple:
    """从 txt 行按指定列（默认第 1 列）提取 ES URL/IP。

    每行按空白/逗号/分号/制表符分隔为多列，返回 (targets, max_cols, target_rows)。
    """
    targets, seen, target_rows = [], set(), {}
    max_cols = 0
    for row_no, line in enumerate(lines, start=1):
        fields = [f for f in re.split(r"[\s,;|\t]+", line) if f]
        max_cols = max(max_cols, len(fields))
        if column > len(fields):
            continue
        t = parse_target(fields[column - 1])
        if t and t["target"] not in seen:
            seen.add(t["target"])
            targets.append(t)
            target_rows[t["target"]] = row_no
    return targets, max_cols, target_rows


def save_batch_context(batch_id: str, ctx: dict) -> None:
    """持久化上传批次上下文（输入类型、原始表格、目标行号映射）。"""
    with open(os.path.join(BATCH_DIR, batch_id + ".json"), "w", encoding="utf-8") as f:
        json.dump(ctx, f, ensure_ascii=False)


def load_batch_context(batch_id: str):
    """读取上传批次上下文，不存在时返回 None。"""
    if not batch_id:
        return None
    path = os.path.join(BATCH_DIR, batch_id + ".json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


def is_es_response(resp) -> bool:
    """粗略判断 HTTP 响应是否来自 Elasticsearch。"""
    if resp.status_code != 200:
        return False
    text = resp.text[:2000]
    return any(k in text for k in ("cluster_name", '"version"', "You Know, for Search", "tagline"))


def _windows_system_proxy() -> str:
    """读取 Windows 系统代理（Internet 选项 -> 局域网设置）中的代理服务器。"""
    try:
        import winreg
    except ImportError:
        return ""
    try:
        key = winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Internet Settings")
        enable, _ = winreg.QueryValueEx(key, "ProxyEnable")
        if not enable:
            return ""
        server, _ = winreg.QueryValueEx(key, "ProxyServer")
        if not server:
            return ""
        # 可能形如 "127.0.0.1:7890" 或 "http=127.0.0.1:7890;socks=127.0.0.1:1080"
        for part in server.split(";"):
            part = part.strip()
            if not part:
                continue
            if "=" in part:
                k, v = part.split("=", 1)
                if k.lower() in ("http", "https"):
                    return v.strip()
            else:
                return part
    except OSError:
        pass
    return ""


def get_system_proxy() -> dict:
    """默认采用系统代理：优先环境变量（HTTP_PROXY/HTTPS_PROXY），
    其次读取 Windows 系统代理设置；两者均未配置时返回 None（直连）。
    """
    http_proxy = os.environ.get("HTTP_PROXY") or os.environ.get("http_proxy")
    if http_proxy:
        https_proxy = os.environ.get("HTTPS_PROXY") or os.environ.get("https_proxy") or http_proxy
        return {"http": http_proxy, "https": https_proxy}
    sp = _windows_system_proxy()
    if sp:
        return {"http": sp, "https": sp}
    return None


_http_session = None


def get_http_session():
    """全局复用连接池，禁用 urllib3 重试，避免不可达目标耗时放大。

    默认采用系统代理（环境变量 / Windows 系统代理），VPN 环境可正常访问目标。
    """
    global _http_session
    if _http_session is None:
        s = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            max_retries=0, pool_connections=MAX_THREADS, pool_maxsize=MAX_THREADS)
        s.mount("http://", adapter)
        s.mount("https://", adapter)
        proxies = get_system_proxy()
        if proxies:
            s.proxies.update(proxies)
            logger.info("已采用系统代理：%s", proxies)
        else:
            logger.info("未检测到系统代理，采用直连方式访问目标")
        s.verify = False   # 跳过证书校验，兼容自签证书的 ES https 目标
        try:
            requests.packages.urllib3.disable_warnings()
        except Exception:
            pass
        _http_session = s
    return _http_session


def _candidate_bases(target: dict) -> list:
    """按目标生成候选检测地址列表（协议/端口自动尝试）。

    规则：显式指定的协议与端口优先；未指定时依次尝试 http/https 的常见 ES 端口，
    仅在连接失败时切换候选，避免重复耗时。
    """
    host = target.get("host", "")
    scheme = target.get("scheme")
    port = target.get("port")
    cands = []
    if port:
        if scheme:
            cands.append((scheme, port))
        else:
            cands.append(("http", port))      # 未写协议时先试 http
            cands.append(("https", port))     # 连接失败再试 https 同端口
    else:
        if scheme == "https":
            cands.append(("https", DEFAULT_PORT))   # ES 默认 9200（TLS）
            cands.append(("https", 443))            # 常见 https 入口端口
        elif scheme == "http":
            cands.append(("http", DEFAULT_PORT))
            cands.append(("https", DEFAULT_PORT))
        else:   # 未写协议：同端口 http/https 都试，再补 https 443
            cands.append(("http", DEFAULT_PORT))
            cands.append(("https", DEFAULT_PORT))
            cands.append(("https", 443))
    out = []
    for s, p in cands:
        if (s, p) not in out:
            out.append((s, p))
    return [f"{s}://{host}:{p}" for s, p in out]


def check_es_weak_password(target: dict, deep: bool = False) -> dict:
    """检测单个 ES 目标是否存在弱口令。target 为 parse_target 返回的 dict。

    deep=True 时使用规则化生成的深度口令组合（爆破式），尝试数量更多、耗时更长。
    """
    if not REQUESTS_OK:
        return {"status": "failure", "weak": False, "detail": "缺少 requests 依赖"}
    tname = target.get("target", "")
    cands = _candidate_bases(target)
    logger.info("检测目标：%s，候选地址：%s，模式：%s",
                tname, cands, "深度" if deep else "常规")

    # 1. 连通性探测：无认证访问。依次尝试候选协议/端口，首个能建立连接的地址
    #    固定用于后续弱口令验证（http(s) 与 9200/443 自动切换，适配 https 目标）
    base_url = None
    for cand in cands:
        try:
            resp = get_http_session().get(cand + "/", timeout=(2, 3))
            if is_es_response(resp):
                logger.info("目标 %s：%s 无需认证即可访问（存在弱口令风险）", tname, cand)
                return {"status": "success", "weak": True, "detail": "未启用认证即可访问（存在弱口令风险）"}
            if resp.status_code >= 500:
                logger.warning("目标 %s：%s 返回 HTTP %s（代理/网关错误），尝试下一候选",
                               tname, cand, resp.status_code)
                continue   # 502/503 等代理或网关错误，尝试下一个协议/端口候选
            base_url = cand   # 已收到目标有效响应（如 401/403），固定该地址
            logger.info("目标 %s：%s 可连接（HTTP %s），固定用于弱口令验证",
                        tname, cand, resp.status_code)
            break
        except requests.exceptions.ConnectionError:
            logger.warning("目标 %s：%s 无法建立连接", tname, cand)
            continue
        except requests.exceptions.RequestException as e:
            logger.warning("目标 %s：%s 请求异常：%s", tname, cand, e)
            continue
    if base_url is None:
        logger.warning("目标 %s：所有候选地址均无法建立连接或响应异常", tname)
        return {"status": "failure", "weak": False, "detail": "无法建立连接或响应异常"}

    # 2. 尝试弱口令组合：常规模式用精选字典，深度模式用规则化爆破组合
    creds = DEEP_CREDENTIALS if deep else WEAK_PASSWORDS
    for user, pwd in creds:
        try:
            resp = get_http_session().get(base_url + "/", auth=(user, pwd), timeout=(2, 3))
            if is_es_response(resp):
                logger.info("目标 %s：命中弱口令 %s / %s", tname, user, pwd)
                return {"status": "success", "weak": True, "detail": f"弱口令：{user} / {pwd}"}
        except requests.exceptions.RequestException:
            continue

    # 3. 能连接但需认证且弱口令均失败 -> 检测成功（未检测到弱口令）
    if deep:
        fail_tip = "未检测到弱口令（已尝试 %d 组深度检测口令）" % len(creds)
    else:
        fail_tip = "未检测到弱口令（常见弱口令均已尝试）"
    try:
        resp = get_http_session().get(base_url + "/", timeout=(2, 3))
        if resp.status_code == 401:
            logger.info("目标 %s：未检测到弱口令（HTTP 401）", tname)
            return {"status": "success", "weak": False, "detail": fail_tip}
        if resp.status_code == 403:
            logger.info("目标 %s：未检测到弱口令（HTTP 403 拒绝访问）", tname)
            return {"status": "success", "weak": False, "detail": "未检测到弱口令（HTTP 403 拒绝访问）"}
        if resp.status_code == 200:
            logger.warning("目标 %s：%s 可连接但响应非 Elasticsearch 服务（HTTP 200），"
                           "可能为管理面板/网页而非 ES REST API", tname, base_url)
            return {"status": "failure", "weak": False,
                    "detail": "目标可连接但响应非 Elasticsearch 服务，请确认协议/端口是否为 ES REST API 地址"}
    except requests.exceptions.RequestException:
        pass

    logger.warning("目标 %s：%s 最终判定为无法建立连接或响应异常", tname, base_url)
    return {"status": "failure", "weak": False, "detail": "无法建立连接或响应异常"}


# ==================== 路由 ====================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload", methods=["POST"])
def upload_file():
    """文件上传：仅保存输入文件上下文，不做 URL 校验。

    xlsx 允许不填写列数（不默认第 1 列），URL 列校验推迟到点击"开始查询"时；
    txt 默认取第 1 列。
    """
    f = request.files.get("file")
    if f is None or not f.filename:
        return jsonify({"code": 1, "msg": "未选择文件"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"code": 1, "msg": "仅支持 .xlsx 或 .txt 文件"}), 400

    column = validate_column(request.form.get("column", ""))   # 允许为空：xlsx 不默认第 1 列

    f.stream.seek(0, os.SEEK_END)
    size = f.stream.tell()
    f.stream.seek(0)
    if size > MAX_FILE_SIZE:
        return jsonify({"code": 1, "msg": "文件大小超过限制（最大 10MB）"}), 400

    save_name = uuid.uuid4().hex + ext
    save_path = os.path.join(UPLOAD_DIR, save_name)
    f.save(save_path)
    batch_id = uuid.uuid4().hex
    try:
        if ext == ".xlsx":
            rows, max_cols = read_xlsx_rows(save_path)
            ctx = {"src": "xlsx", "filename": f.filename, "rows": rows,
                   "max_cols": max_cols, "column": column or 0}
        else:
            lines = read_txt_lines(save_path)
            column = 1 if column is None else column   # txt 默认取第 1 列
            _, max_cols, _ = extract_targets_from_lines(lines, column)
            ctx = {"src": "txt", "filename": f.filename, "lines": lines,
                   "max_cols": max_cols, "column": column}
    except Exception as e:
        safe_remove(save_path)
        return jsonify({"code": 1, "msg": f"文件解析失败：{e}"}), 400
    finally:
        safe_remove(save_path)

    if ext == ".xlsx" and not ctx["rows"]:
        return jsonify({"code": 1, "msg": "表格中没有数据"}), 400
    if ext == ".txt" and not ctx["lines"]:
        return jsonify({"code": 1, "msg": "txt 文件为空或没有有效行"}), 400

    save_batch_context(batch_id, ctx)

    logger.info("文件上传成功：%s（batch=%s，类型=%s，行数=%d，URL列=%s）",
                f.filename, batch_id, ext,
                len(ctx["rows"]) if ext == ".xlsx" else len(ctx["lines"]), ctx["column"])

    return jsonify({
        "code": 0, "msg": "文件已上传",
        "batch_id": batch_id,
        "src": ctx["src"],
        "column": ctx["column"],
        "max_cols": ctx.get("max_cols", 0),
        "line_count": len(ctx["rows"]) if ext == ".xlsx" else len(ctx["lines"]),
        "filename": f.filename,
    })


@app.route("/api/check", methods=["POST"])
def run_check():
    """启动弱口令检测：点击"开始查询"时按当前列数从已上传文件解析并校验 URL，通过后并发检测。"""
    data = request.get_json(silent=True) or {}

    batch_id = str(data.get("batch_id") or "")[:32]
    ctx = load_batch_context(batch_id)
    if not ctx:
        return jsonify({"code": 1, "msg": "未找到已上传的文件，请重新上传"}), 400

    src = ctx.get("src", "txt")
    column = validate_column(data.get("column"))
    deep = bool(data.get("deep"))   # 是否启用深度检测（密码爆破式）

    # 通过 URL 所在列数检查输入文件是否正确输入：列为空 / 超界 / 该列无有效 URL 均提示，
    # 报错时不执行检测、也不删除已上传文件。
    if src == "xlsx":
        if column is None:
            return jsonify({"code": 1, "msg": "请填写 URL 所在列数后再开始查询"}), 400
        max_cols = int(ctx.get("max_cols") or 0)
        if column > max_cols:
            return jsonify({"code": 1,
                            "msg": f"列数（第 {column} 列）超过表格实际列数（共 {max_cols} 列），请检查输入文件是否正确"}), 400
        targets, target_rows = extract_targets_from_rows(ctx.get("rows") or [], column)
        if not targets:
            return jsonify({"code": 1,
                            "msg": f"第 {column} 列中未找到有效的 ES URL/IP 地址，请检查输入文件是否正确或修改列数"}), 400
    else:
        column = 1 if column is None else column   # txt 默认取第 1 列
        lines = ctx.get("lines") or []
        targets, _, target_rows = extract_targets_from_lines(lines, column)
        if not targets:
            return jsonify({"code": 1,
                            "msg": "文件中未找到有效的 ES URL/IP 地址，请检查输入文件是否正确"}), 400

    # 记录本次查询实际使用的列与目标行映射（供结果导出时还原输入表格）
    ctx["column"] = column
    ctx["target_rows"] = target_rows
    save_batch_context(batch_id, ctx)

    logger.info("开始检测：batch=%s，文件名=%s，目标数=%d，URL列=%s，模式=%s",
                batch_id, ctx.get("filename"), len(targets), column,
                "深度" if deep else "常规")

    results = []
    with ThreadPoolExecutor(max_workers=MAX_THREADS) as pool:
        futures = {pool.submit(check_es_weak_password, t, deep): t for t in targets}
        for fut in as_completed(futures):
            t = futures[fut]
            try:
                res = fut.result()
            except Exception:
                res = {"status": "failure", "weak": False, "detail": "检测过程发生异常"}
            results.append({
                "id": uuid.uuid4().hex,
                "batch_id": batch_id,
                "src": src,
                "target": t["target"],
                "ip": t["host"],
                "port": t.get("port") or DEFAULT_PORT,
                "status": res.get("status", "failure"),
                "weak": bool(res.get("weak")),
                "detail": res.get("detail", ""),
                "deep": deep,
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })

    all_results = load_results()
    all_results.extend(results)
    save_results(all_results)
    weak_count = sum(1 for r in results if r.get("weak"))
    conn_ok = sum(1 for r in results if r.get("status") == "success")
    logger.info("检测完成：batch=%s，共 %d 个目标，命中弱口令 %d 个，连接成功 %d 个",
                batch_id, len(results), weak_count, conn_ok)
    batch_info = _aggregate_batches(results)[0] if results else None
    return jsonify({"code": 0, "msg": "检测完成", "results": results, "batch": batch_info})


@app.route("/api/results", methods=["GET"])
def get_results():
    """返回按输入文件（批次）聚合后的检测记录，每输入文件一行。"""
    return jsonify({"code": 0, "results": _aggregate_batches(load_results())})


@app.route("/api/results", methods=["DELETE"])
def clear_results():
    save_results([])
    # 一并清理上传批次上下文，避免残留
    try:
        for fn in os.listdir(BATCH_DIR):
            if fn.endswith(".json"):
                safe_remove(os.path.join(BATCH_DIR, fn))
    except OSError:
        pass
    logger.info("已清空全部检测记录")
    return jsonify({"code": 0, "msg": "已清空所有记录"})


@app.route("/api/results/batch/<batch_id>", methods=["DELETE"])
def delete_batch(batch_id):
    """删除一个输入文件（批次）对应的全部检测记录及其批次上下文。"""
    all_results = load_results()
    before = len(all_results)
    new_results = [r for r in all_results if r.get("batch_id") != batch_id]
    if len(new_results) == before:
        return jsonify({"code": 1, "msg": "记录不存在"}), 404
    save_results(new_results)
    safe_remove(os.path.join(BATCH_DIR, batch_id + ".json"))
    logger.info("已删除批次记录：%s（%d 条）", batch_id, before - len(new_results))
    return jsonify({"code": 0, "msg": "删除成功"})


# ==================== xlsx 结果导出 ====================

RESULT_HEADERS = ["是否存在弱口令", "弱口令详情"]
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _weak_label(item: dict) -> str:
    return "是" if item.get("weak") else "否"


def _aggregate_batches(results: list) -> list:
    """将逐条检测记录按输入文件（批次）聚合为"每文件一行"的记录。

    返回字段：batch_id、filename、status(成功/失败)、reason(失败原因摘要)、
    weak(是否存在弱口令)、detail(详情)、time(检测时间)、total(目标数)。
    """
    groups = {}
    for r in results:
        groups.setdefault(r.get("batch_id") or "legacy", []).append(r)

    out = []
    for batch_id, items in groups.items():
        ctx = load_batch_context(batch_id)
        filename = (ctx or {}).get("filename") or f"batch_{batch_id[:8]}.xlsx"
        total = len(items)
        fail_items = [i for i in items if i.get("status") != "success"]
        weak_items = [i for i in items if i.get("weak")]

        fail_reasons = {}
        for i in fail_items:
            key = i.get("detail") or "检测失败"
            fail_reasons[key] = fail_reasons.get(key, 0) + 1
        reason = "、".join(
            f"{k}×{v}" if v > 1 else k for k, v in fail_reasons.items())

        if weak_items:
            detail = weak_items[0].get("detail") or "存在弱口令"
            if len(weak_items) > 1:
                detail += f"（共 {len(weak_items)} 处）"
        elif fail_items:
            detail = f"共 {total} 个目标，{len(fail_items)} 个检测失败"
        else:
            detail = f"共 {total} 个目标，均未发现弱口令"

        out.append({
            "batch_id": batch_id,
            "filename": filename,
            "status": "success" if not fail_items else "failure",
            "reason": reason if fail_items else "",
            "weak": bool(weak_items),
            "detail": detail,
            "time": items[0].get("time") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total": total,
        })
    return out


def _result_map(results: list) -> dict:
    """按目标字符串建立结果映射（同一目标保留第一条）。"""
    by = {}
    for r in results:
        key = r.get("target") or r.get("ip", "")
        if key and key not in by:
            by[key] = r
    return by


def _batch_rows(ctx: dict, results: list) -> list:
    """生成一个批次的结果表格行（含表头）。

    xlsx 输入：只保留"URL 所在列数"参数指定的那一列，不保留其他列，
    输出为：URL + 是否存在弱口令 + 弱口令详情；
    txt 输入：第一列为 URL，第二列起为结果列。
    """
    if ctx and ctx.get("src") == "xlsx" and ctx.get("rows"):
        rows = ctx["rows"]
        target_rows = ctx.get("target_rows") or {}
        column = int(ctx.get("column") or 1)
        by_target = _result_map(results)
        row_result = {rn: by_target[t] for t, rn in target_rows.items() if t in by_target}
        first_is_data = 1 in target_rows.values()   # 第一行是否就是数据行（无表头）
        out = []
        for i, row in enumerate(rows):
            row_no = i + 1
            vals = ["" if v is None else v for v in row]
            url_val = vals[column - 1] if column <= len(vals) else ""
            if row_no == 1 and not first_is_data:
                out.append([url_val] + list(RESULT_HEADERS))
            else:
                res = row_result.get(row_no)
                if res:
                    out.append([url_val, _weak_label(res), res.get("detail", "")])
                else:
                    out.append([url_val, "", ""])
        return out
    # txt 或未知批次：第一列为 URL
    out = [["URL"] + list(RESULT_HEADERS)]
    for r in results:
        out.append([
            r.get("target") or r.get("ip", ""),
            _weak_label(r),
            r.get("detail", ""),
        ])
    return out


def _rows_to_xlsx(rows: list):
    """将表格行列表写入 xlsx（单个工作表，表头加粗，自动列宽）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "检测结果"
    for row in rows:
        ws.append(row)
    if rows:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        widths = {}
        for row in rows:
            for i, v in enumerate(row):
                l = len(str(v)) if v is not None else 0
                if l > widths.get(i, 0):
                    widths[i] = l
        for i, w in widths.items():
            ws.column_dimensions[get_column_letter(i + 1)].width = min(w + 2, 60)
    return wb


def _send_xlsx(wb, filename: str):
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=filename, mimetype=XLSX_MIME)


@app.route("/api/results/download/batch", methods=["GET"])
def download_batch_results():
    """下载一个输入文件（批次）的检测结果：还原输入表格并追加结果列，始终只输出一个表格。"""
    batch_id = request.args.get("batch_id", "")
    items = [r for r in load_results() if r.get("batch_id") == batch_id]
    if not items:
        return jsonify({"code": 1, "msg": "该次查询暂无检测记录，请先执行查询"}), 400
    logger.info("下载批次结果：batch=%s，记录 %d 条", batch_id, len(items))
    rows = _batch_rows(load_batch_context(batch_id), items)
    wb = _rows_to_xlsx(rows)
    ctx = load_batch_context(batch_id)
    base = os.path.splitext((ctx or {}).get("filename") or "结果")[0]
    filename = f"{base}_检测结果.xlsx"
    return _send_xlsx(wb, filename)


@app.route("/api/results/download/all")
def download_all_results():
    """下载全部检测记录（xlsx，按输入批次还原为单个表格）。"""
    all_results = load_results()
    if not all_results:
        return jsonify({"code": 1, "msg": "暂无检测记录"}), 400
    logger.info("下载全部结果：%d 条记录", len(all_results))

    # 按输入批次分组；同一批次还原输入表格，多批次合并为一个表格
    groups = {}
    for item in all_results:
        groups.setdefault(item.get("batch_id") or "legacy", []).append(item)

    if len(groups) == 1:
        batch_id, items = next(iter(groups.items()))
        rows = _batch_rows(load_batch_context(batch_id), items)
    else:
        max_cols = 0
        for batch_id, items in groups.items():
            r = _batch_rows(load_batch_context(batch_id), items)
            max_cols = max(max_cols, max((len(x) for x in r), default=0))
        rows = []
        first = True
        for batch_id, items in groups.items():
            r = _batch_rows(load_batch_context(batch_id), items)
            # xlsx 无表头（第一行即数据）时批次本身不含表头行；
            # 含表头的批次在合并时仅保留第一批次的表头
            ctx = load_batch_context(batch_id)
            has_header = not (ctx and ctx.get("src") == "xlsx"
                              and 1 in (ctx.get("target_rows") or {}).values())
            for idx, row in enumerate(r):
                if has_header and idx == 0 and not first:
                    continue
                rows.append(list(row) + [""] * (max_cols - len(row)))
            first = False

    wb = _rows_to_xlsx(rows)
    filename = f"es_check_results_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return _send_xlsx(wb, filename)


if __name__ == "__main__":
    # threaded=True：避免单线程开发服务器在处理慢请求/keep-alive 时阻塞其他请求
    logger.info("ES 弱口令检测工具启动，请在浏览器访问 http://127.0.0.1:5000")
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
